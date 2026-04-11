# catalog/views_admin.py
from django.db import transaction
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import load_workbook
from rest_framework.parsers import JSONParser, MultiPartParser
from .serializers_admin import _clean_station


from .models import Product, ProductVariant, Category
from .serializers_admin import ProductAdminSer
from orders.permissions import InGroups  # your existing group-perm
from rest_framework import viewsets
from rest_framework.parsers import JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend

from .models import Product, ProductVariant, Category, ModifierGroup
from .serializers_admin import ProductAdminSer, ModifierGroupSimpleSer
from orders.permissions import InGroups

from .serializers_admin import CategoryAdminSer
from .models import Category

class CategoryAdminViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all().order_by("sort", "name")
    serializer_class = CategoryAdminSer
    permission_classes = [IsAuthenticated, InGroups("manager")]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ["name"]
    ordering_fields = ["id", "name", "sort", "is_active"]
    ordering = ["sort", "name"]


class ProductAdminViewSet(viewsets.ModelViewSet):
    queryset = (
        Product.objects.select_related("category")
        .prefetch_related("variants", "modifier_links__group")
    )
    serializer_class = ProductAdminSer
    permission_classes = [IsAuthenticated, InGroups("manager")]
    # ← FIX 415: accept JSON (and keep multipart for Excel import)
    parser_classes = [JSONParser, MultiPartParser]

    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ["category", "is_active"]
    search_fields = ["name", "kitchen_station"]
    ordering_fields = ["id", "name", "base_price"]
    ordering = ["id"]

    @action(detail=False, methods=["post"])
    def import_excel(self, request):
        """
        form-data:
          file: <xlsx>   (required)
          dry_run: true|false   (optional; default false)
          category_mode: name|id (optional; default 'name')

        Required columns:
          category,name,base_price,kitchen_station,is_active,
          sop_text,sop_audio_url,variant_name,variant_price_delta
        """
        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "No file uploaded."}, status=400)

        dry_run = str(request.data.get("dry_run", "false")).lower() == "true"
        cat_mode = (request.data.get("category_mode") or "name").lower()

        try:
            wb = load_workbook(f, data_only=True)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        except StopIteration:
            return Response({"detail": "Empty sheet"}, status=400)
        except Exception as e:
            return Response({"detail": f"Invalid Excel: {e}"}, status=400)

        required = ["category", "name", "base_price"]
        for col in required:
            if col not in headers:
                return Response({"detail": f"Missing column: {col}"}, status=400)

        idx = {h: headers.index(h) for h in headers}
        created_products = 0
        created_variants = 0
        errors = []

        def val(row, col):
            if col not in idx: return None
            cell = row[idx[col]]
            return getattr(cell, "value", None)

        @transaction.atomic
        def do_import():
            nonlocal created_products, created_variants
            cache = {}
            name_to_cat = {}
            if cat_mode == "name":
                name_to_cat = {c.name.strip().lower(): c for c in Category.objects.all()}

            for r, row in enumerate(ws.iter_rows(min_row=2), start=2):
                try:
                    cat_raw = val(row, "category")
                    name = (val(row, "name") or "").strip()
                    base_price = val(row, "base_price")
                    if not name:
                        raise ValueError("name is required")
                    if base_price is None:
                        raise ValueError("base_price is required")

                    kitchen_station = _clean_station(val(row, "kitchen_station") or "")


                    is_active = val(row, "is_active")
                    sop_text = val(row, "sop_text") or ""
                    sop_audio_url = val(row, "sop_audio_url") or ""
                    v_name = (val(row, "variant_name") or "").strip()
                    v_delta = val(row, "variant_price_delta") or 0

                    # resolve category
                    if cat_mode == "id":
                        try:
                            category = Category.objects.get(pk=int(cat_raw))
                        except Exception:
                            raise ValueError(f"category id not found: {cat_raw}")
                    else:
                        key = str(cat_raw or "").strip().lower()
                        category = name_to_cat.get(key)
                        if not category:
                            raise ValueError(f"category name not found: {cat_raw}")

                    key = (category.id, name)
                    p = cache.get(key)
                    if p is None:
                        p = Product(
                            category=category,
                            name=name,
                            base_price=base_price,
                            kitchen_station=kitchen_station,
                            is_active=bool(is_active) if is_active is not None else True,
                            sop_text=sop_text,
                            sop_audio_url=sop_audio_url,
                        )
                        if not dry_run:
                            p.save()
                        cache[key] = p
                        created_products += 1

                    if v_name:
                        if not dry_run:
                            ProductVariant.objects.create(
                                product=p, name=v_name, price_delta=v_delta
                            )
                        created_variants += 1

                except Exception as e:
                    errors.append({"row": r, "error": str(e)})

            if errors and not dry_run:
                raise transaction.TransactionManagementError("Import has errors; rolled back.")

        if dry_run:
            try:
                with transaction.atomic():
                    do_import()
                    raise transaction.Rollback
            except transaction.Rollback:
                pass
            except Exception:
                pass
        else:
            try:
                do_import()
            except Exception:
                return Response(
                    {"created_products": created_products, "created_variants": created_variants, "errors": errors},
                    status=400,
                )

        return Response(
            {"created_products": created_products, "created_variants": created_variants, "errors": errors}
        )


# class ModifierGroupAdminViewSet(viewsets.ReadOnlyModelViewSet):
#     """
#     Manager can fetch available modifier groups to attach to products.
#     GET /api/admin/modifier-groups/?search=
#     """
#     queryset = ModifierGroup.objects.all().order_by("name")
#     serializer_class = ModifierGroupSimpleSer
#     permission_classes = [IsAuthenticated, InGroups("manager")]
#     filter_backends = [SearchFilter]
#     search_fields = ["name"]

from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend

from .models import ModifierGroup
from .serializers_admin import ModifierGroupAdminSer  # ← use the writeable serializer
from orders.permissions import InGroups

class ModifierGroupAdminViewSet(viewsets.ModelViewSet):
    """
    CRUD for modifier groups (and nested options) for managers.
    """
    queryset = (
        ModifierGroup.objects.all()
        .prefetch_related("options", "product_links__product")
        .order_by("name")
    )
    serializer_class = ModifierGroupAdminSer
    permission_classes = [IsAuthenticated, InGroups("manager")]
    filter_backends = [SearchFilter, OrderingFilter, DjangoFilterBackend]
    search_fields = ["name"]
    ordering_fields = ["id", "name"]
